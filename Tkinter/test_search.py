import openpyxl
import re
import os
from time import sleep
from openpyxl import load_workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils import coordinate_from_string
#from timer import prompt_with_timeout
from test_timer import studentGUI
from reporting import logs
import random
from clockin import timesheet
from tkinter import *

timing = timesheet()

def remove(y):
    
    cell_location=re.sub(r'[\D]','',y)
    return cell_location

def search_Student (x):
    wb=load_workbook('testing.xlsx',read_only = False, data_only = True)
    worksheet= wb.get_sheet_names()
    wb.active
    sheet = wb.get_sheet_by_name('Sheet')
    max_row = sheet.max_row
    insert_name=max_row+1

    max_col=sheet.max_column
    col=get_column_letter(1)# convert column number to letter and use for first column (ID card data)
    col2=get_column_letter(2)# use for second column (Student name)
    col3=get_column_letter(3)# use for third column (Student's email)
    col4=get_column_letter(max_col)#use for third column(occurance)
    #print (col4)
    count =0
    first_time =1
    for row in sheet.iter_rows():
        for cell in row:
            data = cell.value
            if data == str(x):
                cell = str(cell)
                cool=int(remove(cell))
                stop_by= sheet.cell('%s%d' % (col4,cool)).value
                if stop_by == None:
                    stop_by = 0
                #print (stop_by)
                sheet['%s%d' % (col4,cool)] = stop_by+1
                #print (sheet.cell('%s%d' % (col4,cool)).value)
                wb.save('testing.xlsx')
                count+=1
                again=sheet.cell('%s%d' % (col2,cool)).value
                email=sheet.cell('%s%d' % (col3,cool)).value
               
    if count< 1 :
        print ('\a\a\a\a\a\a')
        os.system ('clear')
        #os.system ('echo "New Student"')
        print ("\t\t****** NEW STUDENT ******\n")
        
        root = Tk()
        #search = studentGUI(root)        
        #student_name = search.prompt_with_timeout()
        student_name = studentGUI(root).prompt_with_timeout()

        print("student_name =", student_name)
        #student_name = 'Paul'
        #student_name , student_email = search.prompt_with_timeout
        #student_name = search.prompt_with_timeout()

        print("student_name(from if) =", student_name)
        #again = student_name


        sheet['%s%d' % (col2,insert_name)] =str(student_name)
        sheet['%s%d' % (col,insert_name)] =str(x)
        sheet['%s%d' % (col4,insert_name)] =first_time
        #sheet['%s%d' % (col3,insert_name)] =str(student_email)
        #logs(x,student_name)
        #timing.register(student_name, x)
        #timing.excel(x)
        wb.save('testing.xlsx')
        return "****** NEW STUDENT ******"
    else:
        print ('\a')
        #logs(x,again)
        #timing.register(again, x)
        #timing.excel(x)
        if again == None or  email == None or again == "None" or  email == "None":
            print ('\a\a\a\a\a\a')
            #student_name=input("Enter Name -> ")
            #student_email=input("Enter Student's email -> ")
            #root = Tk()
            #search = studentGUI(root)
            #student_name , student_email = search.prompt_with_timeout
            
            root = Tk()
            #search = studentGUI(root)
            #student_name = search.prompt_with_timeout()
            #search.prompt_with_timeout()
            student_name = studentGUI(root).prompt_with_timeout()
            
            print("student_name(from else) =", student_name)
            sheet['%s%d' % (col2,cool)] = str(student_name)
            #sheet['%s%d' % (col3,cool)] =str(student_email)
            #again = sheet.cell('%s%d' % (col2,cool)).value
            #email = sheet.cell('%s%d' % (col3,cool)).value
            
            #print (x)
            wb.save('testing.xlsx')
        
        print ("\n\tWelcome Back! ",again,"\n")
        return "Welcome Back! " + again